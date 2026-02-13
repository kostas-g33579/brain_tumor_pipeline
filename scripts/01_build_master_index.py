"""
Step 1 — Build Master Index
===========================
Scans both GLI and MEN dataset directories and builds a single master_index.csv
that maps every case to its file paths, label, site/batch info, and split.

This CSV is the single source of truth that drives all downstream pipeline steps.

Usage:
    python 01_build_master_index.py --config ../config/config.yaml
    python 01_build_master_index.py --config ../config/config.yaml --dry-run
"""

import os
import csv
import glob
import argparse
import logging
import yaml
import openpyxl
from pathlib import Path
from datetime import datetime


def load_config(config_path):
    with open(config_path, "r") as f:
        return yaml.safe_load(f)


def setup_logging(cfg):
    os.makedirs(cfg["paths"]["logs_dir"], exist_ok=True)
    log_file = os.path.join(
        cfg["paths"]["logs_dir"],
        f"01_build_master_index_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
    )
    logging.basicConfig(
        level=getattr(logging, cfg["logging"]["level"]),
        format="%(asctime)s [%(levelname)s] %(message)s",
        handlers=[
            logging.StreamHandler(),
            logging.FileHandler(log_file) if cfg["logging"]["save_logs"] else logging.NullHandler()
        ]
    )
    return logging.getLogger(__name__)


def load_gli_mapping(xlsx_path):
    """
    Load GLI mapping XLSX → returns dict: {brats2023_id: site_no}
    """
    mapping = {}
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        case_id = row[0]   # BraTS2023 ID
        site_no = row[7]   # Site No (institution)
        cohort  = row[6]   # Cohort name
        if case_id:
            mapping[case_id] = {
                "site": str(site_no) if site_no else "unknown",
                "cohort": str(cohort) if cohort else "unknown"
            }
    return mapping


def load_men_clinical(xlsx_path):
    """
    Load MEN clinical XLSX → returns dict: {brats_id: {grade, age, sex, split}}
    """
    clinical = {}
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        case_id = row[0]
        clinical[case_id] = {
            "split":  str(row[1]) if row[1] else "unknown",
            "grade":  str(row[2]) if row[2] else "n/a",
            "age":    str(row[3]) if row[3] else "n/a",
            "sex":    str(row[4]) if row[4] else "n/a",
            # Imaging resolution params (useful as ComBat covariates)
            "t1c_slice_thickness": str(row[10]) if row[10] else "n/a",
        }
    return clinical


def find_case_files(case_dir, case_id, modality_suffixes, active_modalities):
    """
    For a given case directory, find file paths for each active modality + seg.
    Returns dict of {modality: filepath or None}
    """
    files = {}
    for mod, enabled in active_modalities.items():
        if not enabled:
            files[mod] = ""
            continue
        suffix = modality_suffixes[mod]
        expected = os.path.join(case_dir, case_id + suffix)
        files[mod] = expected if os.path.exists(expected) else ""

    # Segmentation mask
    seg_suffix = modality_suffixes["seg"]
    seg_path = os.path.join(case_dir, case_id + seg_suffix)
    files["seg"] = seg_path if os.path.exists(seg_path) else ""

    return files


def scan_dataset(dataset_dir, label_value, dataset_name, modality_suffixes,
                 active_modalities, metadata_dict, split_label, logger):
    """
    Scan a dataset directory and return list of case records.
    dataset_dir: path to training or validation folder
    metadata_dict: site/clinical info from xlsx files
    """
    records = []
    if not os.path.exists(dataset_dir):
        logger.warning(f"Directory not found, skipping: {dataset_dir}")
        return records

    # Each case is a subdirectory
    case_dirs = sorted([
        d for d in os.listdir(dataset_dir)
        if os.path.isdir(os.path.join(dataset_dir, d))
    ])

    logger.info(f"Scanning {dataset_name} [{split_label}]: found {len(case_dirs)} case directories")

    missing_files = 0
    for case_id in case_dirs:
        case_dir = os.path.join(dataset_dir, case_id)
        files = find_case_files(case_dir, case_id, modality_suffixes, active_modalities)

        # Check if any active modality file is missing
        active_missing = [
            mod for mod, enabled in active_modalities.items()
            if enabled and files.get(mod) == ""
        ]

        if active_missing:
            missing_files += 1
            logger.warning(f"  MISSING modalities {active_missing} for case: {case_id}")

        # Get metadata
        meta = metadata_dict.get(case_id, {})

        record = {
            "case_id":      case_id,
            "dataset":      dataset_name,
            "split":        split_label,
            "label":        label_value,
            "label_name":   "glioma" if label_value == 0 else "meningioma",
            "site":         meta.get("site", "n/a"),
            "cohort":       meta.get("cohort", "n/a"),
            "grade":        meta.get("grade", "n/a"),
            "age":          meta.get("age", "n/a"),
            "sex":          meta.get("sex", "n/a"),
            "t1c_slice_thickness": meta.get("t1c_slice_thickness", "n/a"),
            "has_seg":      "yes" if files["seg"] != "" else "no",
            "has_missing":  "yes" if active_missing else "no",
            # File paths
            "path_t1":      files.get("t1", ""),
            "path_t1c":     files.get("t1c", ""),
            "path_t2":      files.get("t2", ""),
            "path_flair":   files.get("flair", ""),
            "path_seg":     files.get("seg", ""),
        }
        records.append(record)

    logger.info(f"  → {len(records)} cases indexed | {missing_files} with missing active modalities")
    return records


def print_summary(records, logger):
    """Print a summary of the indexed dataset."""
    total = len(records)
    gli   = sum(1 for r in records if r["label"] == 0)
    men   = sum(1 for r in records if r["label"] == 1)
    train = sum(1 for r in records if r["split"] == "train")
    val   = sum(1 for r in records if r["split"] == "validate")
    miss  = sum(1 for r in records if r["has_missing"] == "yes")
    no_seg = sum(1 for r in records if r["has_seg"] == "no")

    logger.info("=" * 55)
    logger.info("MASTER INDEX SUMMARY")
    logger.info("=" * 55)
    logger.info(f"  Total cases:          {total}")
    logger.info(f"  Glioma (label=0):     {gli}")
    logger.info(f"  Meningioma (label=1): {men}")
    logger.info(f"  Training split:       {train}")
    logger.info(f"  Validation split:     {val}")
    logger.info(f"  Missing modalities:   {miss}")
    logger.info(f"  Missing seg mask:     {no_seg}")
    logger.info("=" * 55)


def main():
    parser = argparse.ArgumentParser(description="Build master index CSV")
    parser.add_argument("--config", required=True, help="Path to config.yaml")
    parser.add_argument("--dry-run", action="store_true",
                        help="Scan and report without writing CSV")
    args = parser.parse_args()

    cfg    = load_config(args.config)
    logger = setup_logging(cfg)
    logger.info("Starting Step 1: Build Master Index")

    paths     = cfg["paths"]
    mods      = cfg["modalities"]
    suffixes  = cfg["modality_suffixes"]
    labels    = cfg["labels"]

    # ----------------------------------------------------------
    # Load metadata from XLSX files
    # ----------------------------------------------------------
    logger.info("Loading GLI mapping XLSX...")
    gli_meta = load_gli_mapping(paths["gli_mapping_xlsx"])
    logger.info(f"  → {len(gli_meta)} GLI cases in mapping file")

    logger.info("Loading MEN clinical XLSX...")
    men_meta = load_men_clinical(paths["men_clinical_xlsx"])
    logger.info(f"  → {len(men_meta)} MEN cases in clinical file")

    # ----------------------------------------------------------
    # Scan datasets
    # ----------------------------------------------------------
    all_records = []

    # GLI Training
    all_records += scan_dataset(
        paths["gli_training_dir"], labels["glioma"], "GLI",
        suffixes, mods, gli_meta, "train", logger
    )

    # MEN Training
    all_records += scan_dataset(
        paths["men_training_dir"], labels["meningioma"], "MEN",
        suffixes, mods, men_meta, "train", logger
    )

    # Validation sets (if enabled)
    if cfg.get("include_validation_set", False):
        all_records += scan_dataset(
            paths["gli_validation_dir"], labels["glioma"], "GLI",
            suffixes, mods, gli_meta, "validate", logger
        )
        all_records += scan_dataset(
            paths["men_validation_dir"], labels["meningioma"], "MEN",
            suffixes, mods, men_meta, "validate", logger
        )

    # ----------------------------------------------------------
    # Print summary
    # ----------------------------------------------------------
    print_summary(all_records, logger)

    if args.dry_run:
        logger.info("Dry run — no CSV written.")
        return

    # ----------------------------------------------------------
    # Write master CSV
    # ----------------------------------------------------------
    os.makedirs(os.path.dirname(paths["master_csv"]), exist_ok=True)
    fieldnames = list(all_records[0].keys()) if all_records else []

    with open(paths["master_csv"], "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(all_records)

    logger.info(f"Master index written to: {paths['master_csv']}")
    logger.info("Step 1 complete.")


if __name__ == "__main__":
    main()
